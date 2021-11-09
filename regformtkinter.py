#Simple Registration form using openpyxl and tkinter
from openpyxl import *
from tkinter import *

wb = load_workbook('existing excel file path')

sheet = wb.active

def excel():

    #resize the width of columns in excel
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    #write given data to an excel spreadsheet
    sheet.cell(row = 1, column = 1).value = "Name"
    sheet.cell(row = 1, column = 2).value = "Course"
    sheet.cell(row = 1, column = 3).value = "Semester"
    sheet.cell(row = 1, column = 4).value = "Form Number"
    sheet.cell(row = 1, column = 5).value = "Contact Number"
    sheet.cell(row = 1, column = 6).value = "Email ID"
    sheet.cell(row = 1, column = 7).value = "Address"


#function to set focus (cursor)
def focus1(event):
    course_field.focus_set()

#Function to set focus
def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()


#function for clearing the contents of text entry boxes
def clear():

    #clear the content of text entry box
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


#Function to take data from GUI window and write to an excel file
def insert():

    #if user not fill any entry then print "empty input"
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and 
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written in an excel sheet 
        current_row = sheet.max_row
        current_column = sheet.max_column

        #get method returns current text as string 
        #which we write into excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        #save the file
        wb.save('file path')

        name_field.focus_set()

        clear()


if __name__ == "__main__":

    root = Tk()

    root.configure(background= 'light green')

    root.title("registration form")

    root.geometry("500x300")

    excel()

    heading = Label(root, text= "Form", bg= "light green")

    name = Label(root, text = "Name", bg = "light green")

    course = Label(root, text = "Course", bg= "light green")

    sem = Label(root, text= "Semester", bg="light green")

    form_no = Label(root, text= "Form No.", bg="light green")

    contact_no = Label(root, text="Contact No.", bg="light green")

    email_id = Label(root, text="Email id", bg="light green")

    address = Label(root, text="Address", bg = "light green")

    #grid method is used for placing the widgets at respective positions
    #in table like structure.
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)

    #create a text entry box for typing the info
    name_field = Entry(root)
    course_field= Entry(root)
    sem_field= Entry(root)
    form_no_field= Entry(root)
    contact_no_field= Entry(root)
    email_id_field= Entry(root)
    address_field= Entry(root)

    #bind method of widget is used for the binding the function with events
    #whenever the enter key is pressed then call the focus1 function
    name_field.bind("<Return>", focus1)

    #whenever the enter key is pressed then call the focus2 function
    course_field.bind("<Return>", focus2)

    #enter key is pressed then call the focus3 function
    sem_field.bind("<Return>", focus3)

    #enter key is pressed then call the focus4 function
    form_no_field.bind("<Return>", focus4)

    #enter key is pressed then call the focus5 function
    contact_no_field.bind("<Return>", focus5)

    #enter key is pressed then call the focus6 function
    email_id_field.bind("<Return>", focus6)

    #grid method is used for placing the widgets at respective positions
    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1 , ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    #call excel function
    excel()

    #create a Submit button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                            bg="Red, command=insert")
    submit.grid(row=8, column= 1)

    #start the GUI
    root.mainloop()

        